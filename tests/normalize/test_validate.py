from normalize.validate import validate_rows


def _row(price=10.0, name="Product A", ean="1234567890123", sheet="S1"):
    return {
        "price": price,
        "product_name": name,
        "ean": ean,
        "supplier": "TestCo",
        "source_row": 0,
        "source_sheet": sheet,
        "source_file": "test.xlsx",
        "quantity": None,
    }


def test_no_summary_when_all_valid():
    rows = [_row(), _row(ean="9780000000002")]
    _, warnings = validate_rows(rows)
    assert not any("[SKIP SUMMARY]" in w for w in warnings)


def test_summary_counts_no_price():
    rows = [_row(price=None), _row(price=0.0), _row()]
    _, warnings = validate_rows(rows)
    summary = next(w for w in warnings if "[SKIP SUMMARY]" in w)
    assert "2 no_price" in summary


def test_summary_counts_no_name():
    rows = [_row(name=None), _row(name=""), _row()]
    _, warnings = validate_rows(rows)
    summary = next(w for w in warnings if "[SKIP SUMMARY]" in w)
    assert "2 no_name" in summary


def test_summary_counts_duplicate_ean():
    rows = [_row(), _row()]  # same EAN "1234567890123" twice
    _, warnings = validate_rows(rows)
    summary = next(w for w in warnings if "[SKIP SUMMARY]" in w)
    assert "1 duplicate_ean" in summary


def test_summary_shows_only_nonzero_counts():
    rows = [_row(price=None), _row()]
    _, warnings = validate_rows(rows)
    summary = next(w for w in warnings if "[SKIP SUMMARY]" in w)
    assert "no_name" not in summary
    assert "duplicate_ean" not in summary


def test_summary_total_count():
    rows = [_row(price=None), _row(name=None), _row()]
    _, warnings = validate_rows(rows)
    summary = next(w for w in warnings if "[SKIP SUMMARY]" in w)
    assert "2 skipped" in summary


import pytest
from normalize import ingest
from normalize.exceptions import NormalizeError
from pathlib import Path


def test_zero_yield_warning_in_ingest():
    """Sheet 1 has wrong column names (→ [ZERO YIELD]); Sheet 2 has valid data so ingest() succeeds."""
    import tempfile
    import openpyxl
    wb = openpyxl.Workbook()

    # Sheet 1: wrong column headers — mapper returns rows with all-None fields
    # → validate_rows produces 0 products → [ZERO YIELD] warning is emitted
    ws1 = wb.active
    ws1.title = "WrongHeaders"
    ws1.append(["Code", "Label", "Cost"])  # wrong column names
    ws1.append(["12345", "Item A", "5.00"])
    ws1.append(["67890", "Item B", "3.50"])

    # Sheet 2: correct column headers with valid data → ingest() succeeds overall.
    # Need enough valid rows so total yield >= 50% (default min_yield):
    # Sheet 1 has 2 data rows (all invalid) + Sheet 2 has 4 data rows (all valid) = 4/6 = 67% OK
    ws2 = wb.create_sheet("ValidData")
    ws2.append(["EAN", "Name", "Price"])
    ws2.append(["1234567890123", "Prod A", "9.99"])
    ws2.append(["2345678901234", "Prod B", "4.50"])
    ws2.append(["3456789012345", "Prod C", "2.00"])
    ws2.append(["4567890123456", "Prod D", "6.75"])

    with tempfile.NamedTemporaryFile(suffix="_test_offer.xlsx", delete=False) as f:
        tmp_path = f.name
    wb.save(tmp_path)

    try:
        products, warnings = ingest(tmp_path)
        assert len(products) >= 1
        assert any("[ZERO YIELD]" in w for w in warnings)
    finally:
        import os
        try:
            os.unlink(tmp_path)
        except OSError:
            pass  # Windows may hold a file lock; cleanup is best-effort


def test_skip_summary_appears_on_partial_yield():
    """When some rows are skipped (e.g. price=0), a [SKIP SUMMARY] warning must appear."""
    import tempfile
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["EAN", "Name", "Price"])   # correct columns per _test_supplier.yaml
    ws.append(["1234567890123", "Prod A", "9.99"])
    ws.append(["9999999999999", "Prod B", "0"])  # price=0 → skipped
    with tempfile.NamedTemporaryFile(suffix="_test_offer.xlsx", delete=False) as f:
        tmp_path = f.name
    wb.save(tmp_path)

    try:
        products, warnings = ingest(tmp_path)
        assert len(products) == 1
        assert any("[SKIP SUMMARY]" in w for w in warnings)
    finally:
        import os
        try:
            os.unlink(tmp_path)
        except OSError:
            pass  # Windows may hold a file lock; cleanup is best-effort


def test_zero_yield_warning_message():
    """When a sheet produces 0 products due to wrong headers, a [ZERO YIELD] warning must appear."""
    import tempfile
    import openpyxl
    wb = openpyxl.Workbook()

    # Sheet 1: wrong column names → all rows map to None fields → 0 products → [ZERO YIELD]
    ws1 = wb.active
    ws1.title = "BadSheet"
    ws1.append(["Code", "Label", "Cost"])
    ws1.append(["12345", "Item X", "7.50"])

    # Sheet 2: correct columns → valid products so ingest() does not raise.
    # 2 valid rows + 1 bad row = 2/3 ≈ 67% yield — safely above the 50% min_yield threshold.
    ws2 = wb.create_sheet("GoodSheet")
    ws2.append(["EAN", "Name", "Price"])
    ws2.append(["1234567890123", "Prod A", "9.99"])
    ws2.append(["2345678901234", "Prod B", "4.50"])

    with tempfile.NamedTemporaryFile(suffix="_test_offer.xlsx", delete=False) as f:
        tmp_path = f.name
    wb.save(tmp_path)

    try:
        products, warnings = ingest(tmp_path)
        assert len(products) >= 1
        assert any("[ZERO YIELD]" in w for w in warnings)
    finally:
        import os
        try:
            os.unlink(tmp_path)
        except OSError:
            pass  # Windows may hold a file lock; cleanup is best-effort
